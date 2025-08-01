from openpyxl import load_workbook
from pathlib import Path, PureWindowsPath, WindowsPath

file_path = "/home/wilson/workspace/etl_data_fundation/data/COE_GOLD_JUMP.xlsx"

title= 0
rw_path=1
xml_file=2
gpkg_file=3
sld_file=4

origen = "M:\\COE_Digital\\coe_digital_data\\silver_data\\RESTRICTED\\"
destino = "C:\\Temp\\silver_data\\RESTRICTED\\"


collection = []
workbook = load_workbook(filename=file_path, read_only=True)
sheet = workbook.active
for row in sheet.iter_rows(min_row=2, values_only=True):
    rec={"title": row[title],
         "path": row[rw_path], 
         "xml_file": row[xml_file], 
         "gpkg_file": row[gpkg_file], 
         "sld_file": row[sld_file]}
    
    path = str(rec["path"]).replace(origen, destino)

    dir_path = Path(PureWindowsPath(path))
    wp = WindowsPath(dir_path)

    collection.append(rec)
print()
