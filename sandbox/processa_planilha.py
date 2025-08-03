from openpyxl import load_workbook
from pathlib import Path, PureWindowsPath, WindowsPath
import json
import requests

file_path = "/home/wilson/dev/etl_data_fundation/data/COE_GOLD_JUMP.xlsx"

ntitle= 0
nrw_path=1
nxml_file=2
ngpkg_file=3
nsld_file=4

origen = "M:\\COE_Digital\\coe_digital_data\\silver_data\\RESTRICTED\\"
destino = "/home/wilson/silver_data/RESTRICTED/"


collection = []
workbook = load_workbook(filename=file_path, read_only=True)
sheet = workbook.active
with open("upload_log.txt", "w") as log_file:
    records =list(sheet.iter_rows(min_row=2, values_only=True))
    for row in records:
        try:

            path = str(row[nrw_path]).replace(origen, destino).replace("\\", "/")

            title:str = str(row[ntitle])
            xml:str = str(row[nxml_file])
            gpkg:str = str(row[ngpkg_file])
            sld:str = str(row[nsld_file])


            xml_path = Path(path, xml).as_posix()
            gpkg_path = Path(path, gpkg).as_posix()
            sld_path = Path(path, sld).as_posix()

            url = "http://carbono.iocasta.com.br:8201/uploadfiles/"
            #url = "http://127.0.0.1:8000/uploadfiles/"


            payload = {}
            files_payload = [
                ('files', (xml, open(xml_path, 'rb'))),
                ('files', (gpkg, open(gpkg_path, 'rb'))),
                ('files', (sld, open(sld_path, 'rb')))
            ]
            headers = {
            'accept': 'application/json'
            }

            response = requests.request("POST", 
                                        url, 
                                        headers=headers, 
                                        data=payload, 
                                        files=files_payload)
            if response.status_code != 200:
                print(title, response.text)
                continue
            else:
                print(f"Uploaded: {title}")
        except Exception as e:
            print(f"Error processing row {title}")
            continue